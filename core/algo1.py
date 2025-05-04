import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os
import re
import random
import time
import concurrent.futures
from functools import lru_cache
from sklearn.preprocessing import MinMaxScaler
from deap import base, creator, tools, algorithms
POPULATION_SIZE = 100
N_GENERATIONS = 50
CROSSOVER_PROB = 0.7
MUTATION_PROB = 0.2
TOURNAMENT_SIZE = 3
@lru_cache(maxsize=1024)  
def normaliser_horaire(horaire):
    if not isinstance(horaire, str):
        return str(horaire).strip()
    pattern = r'(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})'
    match = re.match(pattern, horaire)
    
    if match:
        h1, m1, h2, m2 = match.groups()
        return f"{int(h1):02d}:{m1} - {int(h2):02d}:{m2}"
    
    return horaire.strip()
class DataManager:
    def __init__(self, dossier="resultats_surveillance"):
        self.dossier = dossier
        self.enseignant_seance = None
        self.seances_par_creneau = None
        self.enseignants_charges = None
        self.toutes_dates = None
        self.dates_converties = {}  
        self.enseignants_idx = {}  
        self.creneaux_idx = {}  
        
    def lire_donnees(self):
        print(f"Lecture des fichiers Excel depuis le dossier '{self.dossier}'...")
        
        self.enseignant_seance = pd.read_excel(os.path.join(self.dossier, 'enseignant_seances.xlsx'))
        self.seances_par_creneau = pd.read_excel(os.path.join(self.dossier, 'seances_par_creneau.xlsx'))
        self.enseignants_charges = pd.read_excel(os.path.join(self.dossier, 'enseignants_charges.xlsx'))
        if 'horaire' in self.enseignant_seance.columns:
            self.enseignant_seance['horaire'] = self.enseignant_seance['horaire'].apply(normaliser_horaire)
        
        if 'horaire' in self.seances_par_creneau.columns:
            self.seances_par_creneau['horaire'] = self.seances_par_creneau['horaire'].apply(normaliser_horaire)
        self.toutes_dates = sorted(set(self.seances_par_creneau['date'].apply(self.convertir_en_datetime)))
        self.enseignants_idx = {row['Nom Et Prénom Enseignant']: i for i, (_, row) in enumerate(self.enseignants_charges.iterrows())}
        self.creneaux_idx = {}
        idx = 0
        for _, row in self.seances_par_creneau.iterrows():
            date = row['date']
            horaire = normaliser_horaire(row['horaire'])
            key = (date, horaire)
            if key not in self.creneaux_idx:
                self.creneaux_idx[key] = idx
                idx += 1
        print(f"Données chargées: {len(self.enseignant_seance)} sessions d'enseignants, "
              f"{len(self.seances_par_creneau)} créneaux, {len(self.enseignants_charges)} enseignants")
        
        return self
    def convertir_en_datetime(self, date):
        if date in self.dates_converties:
            return self.dates_converties[date]
        
        if isinstance(date, datetime):
            result = date
        elif isinstance(date, str):
            try:
                result = datetime.strptime(date, '%Y-%m-%d')
            except ValueError:
                try:
                    result = datetime.strptime(date, '%d/%m/%Y')
                except ValueError:
                    result = pd.to_datetime(date).to_pydatetime()
        else:
            result = pd.to_datetime(date).to_pydatetime()
        self.dates_converties[date] = result
        return result
    def obtenir_seances_enseignants(self):
        seances_enseignants = {}
        for _, row in self.enseignant_seance.iterrows():
            date = row['date']
            horaire = normaliser_horaire(row['horaire'])
            enseignant = row['enseignant']
            
            if enseignant not in seances_enseignants:
                seances_enseignants[enseignant] = set()
            
            seances_enseignants[enseignant].add((date, horaire))
        
        return seances_enseignants
    def obtenir_dates_seances_enseignants(self):
        dates_seances = {}
        for _, row in self.enseignant_seance.iterrows():
            enseignant = row['enseignant']
            date = row['date']
            
            if enseignant not in dates_seances:
                dates_seances[enseignant] = set()
            
            dates_seances[enseignant].add(date)
        
        return dates_seances
    def obtenir_charges_obligatoires(self):
        return {row['Nom Et Prénom Enseignant']: row['charge obligatoire'] 
                for _, row in self.enseignants_charges.iterrows()}
    def obtenir_info_creneaux(self):
        creneaux_info = {}
        for _, row in self.seances_par_creneau.iterrows():
            date = row['date']
            horaire = normaliser_horaire(row['horaire'])
            nb_salles = row['nombre_salles']
            
            creneaux_info[(date, horaire)] = {
                'nb_salles': nb_salles,
                'min_surveillants': 2 * nb_salles,
                'max_surveillants': 3 * nb_salles
            }
        
        return creneaux_info
class GeneticOptimizer:
    def __init__(self, data_manager):
        self.data_manager = data_manager
        self.enseignants = list(data_manager.enseignants_charges['Nom Et Prénom Enseignant'])
        self.creneaux = [(date, horaire) for (date, horaire), _ in data_manager.creneaux_idx.items()]
        self.charges_obligatoires = data_manager.obtenir_charges_obligatoires()
        self.dates_seances_enseignants = data_manager.obtenir_dates_seances_enseignants()
        self.creneaux_info = data_manager.obtenir_info_creneaux()
        self.poids_neuronaux = self._initialiser_poids_neuronaux()
        self._configurer_algorithme_genetique()
    
    def _initialiser_poids_neuronaux(self):
        """Initialise une matrice de poids pour favoriser certaines affectations"""
        n_dates = len(self.data_manager.toutes_dates)
        poids = np.ones((len(self.enseignants), n_dates))
        for i, enseignant in enumerate(self.enseignants):
            if enseignant in self.dates_seances_enseignants:
                dates_seance = self.dates_seances_enseignants[enseignant]
                for date in dates_seance:
                    try:
                        date_idx = self.data_manager.toutes_dates.index(self.data_manager.convertir_en_datetime(date))
                        poids[i, date_idx] = 3.0  
                        if date_idx > 0:
                            poids[i, date_idx-1] = 2.0
                        if date_idx < n_dates - 1:
                            poids[i, date_idx+1] = 2.0
                    except ValueError:
                        continue
        scaler = MinMaxScaler()
        poids_normalized = scaler.fit_transform(poids)
        
        return poids_normalized
    
    def _configurer_algorithme_genetique(self):
        """Configure les composants de l'algorithme génétique avec DEAP"""
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", list, fitness=creator.FitnessMax)
        
        self.toolbox = base.Toolbox()
        self.toolbox.register("map", self._map_parallel)
        self.toolbox.register("attr_bool", self._generate_individual)
        self.toolbox.register("individual", tools.initIterate, creator.Individual, self.toolbox.attr_bool)
        self.toolbox.register("population", tools.initRepeat, list, self.toolbox.individual)
        self.toolbox.register("evaluate", self._evaluer_solution)
        self.toolbox.register("mate", tools.cxTwoPoint)
        self.toolbox.register("mutate", self._mutate)
        self.toolbox.register("select", tools.selTournament, tournsize=TOURNAMENT_SIZE)
    
    def _map_parallel(self, func, seq):
        """Exécute func sur chaque élément de seq en parallèle"""
        with concurrent.futures.ProcessPoolExecutor(max_workers=os.cpu_count()) as executor:
            return list(executor.map(func, seq))
    
    def _generate_individual(self):
        """Génère une solution initiale de répartition des surveillances"""
        n_enseignants = len(self.enseignants)
        n_creneaux = len(self.creneaux)
        solution = np.zeros((n_enseignants, n_creneaux), dtype=int)
        for i, enseignant in enumerate(self.enseignants):
            charge = self.charges_obligatoires.get(enseignant, 0)
            dates_seances = self.dates_seances_enseignants.get(enseignant, set())
            creneaux_possibles = []
            for j, (date, horaire) in enumerate(self.creneaux):
                if date in dates_seances:
                    creneaux_possibles.append(j)
            if len(creneaux_possibles) < charge:
                autres_creneaux = [j for j in range(n_creneaux) if j not in creneaux_possibles]
                random.shuffle(autres_creneaux)
                creneaux_possibles.extend(autres_creneaux[:charge - len(creneaux_possibles)])
            selected = random.sample(creneaux_possibles, min(charge, len(creneaux_possibles)))
            for j in selected:
                solution[i, j] = 1
        
        return solution.flatten().tolist()
    
    def _reshape_solution(self, solution):
        """Convertit la solution linéaire en matrice 2D"""
        n_enseignants = len(self.enseignants)
        n_creneaux = len(self.creneaux)
        return np.array(solution).reshape(n_enseignants, n_creneaux)
    
    def _evaluer_solution(self, solution):
        """Évalue la qualité d'une solution (fitness)"""
        matrice = self._reshape_solution(solution)
        n_enseignants = len(self.enseignants)
        n_creneaux = len(self.creneaux)
        
        score = 0.0
        for j in range(n_creneaux):
            date, horaire = self.creneaux[j]
            nb_surveillants = sum(matrice[i, j] for i in range(n_enseignants))
            info = self.creneaux_info.get((date, horaire), {'min_surveillants': 0, 'max_surveillants': 0})
            if nb_surveillants < info['min_surveillants']:
                score -= 10 * (info['min_surveillants'] - nb_surveillants)
            elif nb_surveillants > info['max_surveillants']:
                score -= 5 * (nb_surveillants - info['max_surveillants'])
        for i, enseignant in enumerate(self.enseignants):
            charge_attribuee = sum(matrice[i, :])
            charge_obligatoire = self.charges_obligatoires.get(enseignant, 0)
            if charge_attribuee > charge_obligatoire:
                score -= 20 * (charge_attribuee - charge_obligatoire)
            elif charge_attribuee < charge_obligatoire:
                score -= 5 * (charge_obligatoire - charge_attribuee)
        for i, enseignant in enumerate(self.enseignants):
            dates_attribuees = set()
            for j in range(n_creneaux):
                if matrice[i, j] == 1:
                    date, _ = self.creneaux[j]
                    dates_attribuees.add(date)
            dates_dt = sorted([self.data_manager.convertir_en_datetime(d) for d in dates_attribuees])
            consecutive = True
            for k in range(len(dates_dt) - 1):
                if (dates_dt[k+1] - dates_dt[k]).days != 1:
                    consecutive = False
                    break
            if consecutive and len(dates_dt) > 1:
                score += 15
            elif len(dates_dt) > 3:  
                score -= 10
            dates_seances = self.dates_seances_enseignants.get(enseignant, set())
            for date in dates_seances:
                if date in dates_attribuees:
                    score += 5
        for i, enseignant in enumerate(self.enseignants):
            for j, (date, _) in enumerate(self.creneaux):
                if matrice[i, j] == 1:
                    try:
                        date_idx = self.data_manager.toutes_dates.index(self.data_manager.convertir_en_datetime(date))
                        score += self.poids_neuronaux[i, date_idx] * 10
                    except ValueError:
                        continue
        
        return (score,)
    
    def _mutate(self, individual, indpb=0.05):
        """Mutation adaptée pour notre problème"""
        matrice = self._reshape_solution(individual)
        n_enseignants = len(self.enseignants)
        n_creneaux = len(self.creneaux)
        
        for i in range(n_enseignants):
            if random.random() < indpb:
                assignes = [j for j in range(n_creneaux) if matrice[i, j] == 1]
                non_assignes = [j for j in range(n_creneaux) if matrice[i, j] == 0]
                
                if assignes and non_assignes:
                    j_retire = random.choice(assignes)
                    j_ajoute = random.choice(non_assignes)
                    
                    matrice[i, j_retire] = 0
                    matrice[i, j_ajoute] = 1
        for idx, val in enumerate(matrice.flatten()):
            individual[idx] = val
        
        return (individual,)
    
    def optimiser(self):
        """Exécute l'algorithme génétique pour trouver une solution optimale"""
        print("Initialisation de l'algorithme génétique...")
        start_time = time.time()
        pop = self.toolbox.population(n=POPULATION_SIZE)
        fitnesses = list(map(self.toolbox.evaluate, pop))
        for ind, fit in zip(pop, fitnesses):
            ind.fitness.values = fit
        
        print(f"Évaluation initiale effectuée en {time.time() - start_time:.2f} secondes")
        stats = tools.Statistics(lambda ind: ind.fitness.values)
        stats.register("avg", np.mean)
        stats.register("min", np.min)
        stats.register("max", np.max)
        print(f"Exécution de l'algorithme génétique sur {N_GENERATIONS} générations...")
        pop, logbook = algorithms.eaSimple(
            pop, self.toolbox,
            cxpb=CROSSOVER_PROB, mutpb=MUTATION_PROB, 
            ngen=N_GENERATIONS, stats=stats, verbose=True
        )
        best_ind = tools.selBest(pop, 1)[0]
        best_solution = self._reshape_solution(best_ind)
        
        print(f"Optimisation terminée en {time.time() - start_time:.2f} secondes")
        
        return best_solution, logbook
class ResultManager:
    def __init__(self, data_manager, solution):
        self.data_manager = data_manager
        self.solution = solution
        self.enseignants = list(data_manager.enseignants_charges['Nom Et Prénom Enseignant'])
        self.creneaux = [(date, horaire) for (date, horaire), _ in data_manager.creneaux_idx.items()]
        
    def extraire_surveillances(self):
        """Extrait la structure de surveillances à partir de la solution"""
        surveillances = {creneau: [] for creneau in self.creneaux}
        for i, enseignant in enumerate(self.enseignants):
            for j, creneau in enumerate(self.creneaux):
                if self.solution[i, j] == 1:
                    surveillances[creneau].append(enseignant)
        
        return surveillances
    
    def extraire_creneaux_par_date(self):
        """Extrait les créneaux de surveillance par date pour chaque enseignant"""
        creneaux_par_date = {enseignant: {} for enseignant in self.enseignants}
        
        for i, enseignant in enumerate(self.enseignants):
            for j, (date, horaire) in enumerate(self.creneaux):
                if self.solution[i, j] == 1:
                    if date not in creneaux_par_date[enseignant]:
                        creneaux_par_date[enseignant][date] = []
                    creneaux_par_date[enseignant][date].append(horaire)
        
        return creneaux_par_date
    
    def extraire_dates_surveillance(self):
        """Extrait les dates de surveillance pour chaque enseignant"""
        dates_surveillance = {enseignant: set() for enseignant in self.enseignants}
        
        for i, enseignant in enumerate(self.enseignants):
            for j, (date, _) in enumerate(self.creneaux):
                if self.solution[i, j] == 1:
                    dates_surveillance[enseignant].add(date)
        
        return dates_surveillance
    
    def generer_excel(self):
        """Génère le fichier Excel de planning de surveillance"""
        creneaux_par_date = self.extraire_creneaux_par_date()
        dossier_sortie = "resultats_surveillance"
        if not os.path.exists(dossier_sortie):
            os.makedirs(dossier_sortie)
        seances_enseignants = self.data_manager.obtenir_seances_enseignants()
        dates_uniques = sorted(set([date for enseignant_dates in creneaux_par_date.values() 
                                for date in enseignant_dates.keys()]))
        dates_uniques = sorted(dates_uniques, key=lambda d: self.data_manager.convertir_en_datetime(d))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planning de surveillance"
        ws.cell(row=1, column=1, value="Enseignant")
        horaires_uniques_par_date = {}
        for date in dates_uniques:
            horaires_pour_date = set()
            for enseignant, dates_info in creneaux_par_date.items():
                if date in dates_info:
                    for horaire in dates_info[date]:
                        horaire_norm = normaliser_horaire(horaire)
                        horaires_pour_date.add(horaire_norm)
            horaires_uniques_par_date[date] = sorted(horaires_pour_date)
        col = 2
        for date in dates_uniques:
            date_display = date.strftime('%Y-%m-%d') if isinstance(date, datetime) else str(date)
            ws.cell(row=1, column=col, value=date_display)
            
            horaires_date = horaires_uniques_par_date[date]
            if len(horaires_date) > 0:
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(horaires_date) - 1)
            for i, horaire in enumerate(horaires_date):
                ws.cell(row=2, column=col + i, value=horaire)
            
            col += len(horaires_date)
        row = 3
        for enseignant in self.enseignants:
            ws.cell(row=row, column=1, value=enseignant)
            
            col = 2
            for date in dates_uniques:
                horaires_date = horaires_uniques_par_date[date]
                
                for horaire_unique in horaires_date:
                    a_cours = False
                    if enseignant in seances_enseignants and (date, horaire_unique) in seances_enseignants[enseignant]:
                        a_cours = True
                    a_surveillance = False
                    if date in creneaux_par_date[enseignant]:
                        horaires_normalises = [normaliser_horaire(h) for h in creneaux_par_date[enseignant][date]]
                        if horaire_unique in horaires_normalises:
                            a_surveillance = True
                    valeur = ""
                    if a_cours and a_surveillance:
                        valeur = "R/S"  
                    elif a_cours:
                        valeur = "R"    
                    elif a_surveillance:
                        valeur = "S"    
                    
                    if valeur:
                        ws.cell(row=row, column=col, value=valeur)
                    
                    col += 1
            
            row += 1
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=1, max_row=row-1, min_col=1, max_col=col-1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
        fichier_excel = os.path.join(dossier_sortie, "planning_surveillance_optimise.xlsx")
        wb.save(fichier_excel)
        
        return fichier_excel
def sont_consecutives(dates_set, data_manager):
    if len(dates_set) <= 1:
        return True
    dates_dt = [data_manager.convertir_en_datetime(date) for date in dates_set]
    dates_triees = sorted(dates_dt)
    for i in range(len(dates_triees) - 1):
        diff = (dates_triees[i+1] - dates_triees[i]).days
        if diff != 1:
            return False
    
    return True
def verifier_contraintes(surveillances, creneaux_par_date, dates_surveillance, data_manager):
    print("\nVérification des contraintes:")
    
    enseignant_seance = data_manager.enseignant_seance
    enseignants_charges = data_manager.enseignants_charges
    seances_par_creneau = data_manager.seances_par_creneau
    print("1. Enseignant surveille le jour de sa matière:")
    for _, row in enseignant_seance.iterrows():
        date = row['date']
        enseignant = row['enseignant']
        if enseignant in creneaux_par_date:  
            if date not in creneaux_par_date[enseignant]:
                print(f"   ❌ Contrainte non respectée: {enseignant} ne surveille pas le {date}")
            else:
                print(f"   ✅ {enseignant} surveille bien le {date}")
        else:
            print(f"   ⚠️ {enseignant} n'est pas dans la liste des enseignants chargés")
    print("\n2. Répartition sur 3 jours consécutifs maximum:")
    for enseignant, dates in dates_surveillance.items():
        if len(dates) > 3:
            if not sont_consecutives(dates, data_manager):
                print(f"   ❌ Contrainte non respectée: {enseignant} a {len(dates)} jours non consécutifs")
                dates_dt = sorted([data_manager.convertir_en_datetime(d) for d in dates])
                print(f"      Dates: {[d.strftime('%Y-%m-%d') for d in dates_dt]}")
            else:
                print(f"   ✅ {enseignant} a {len(dates)} jours strictement consécutifs")
        else:
            print(f"   ✅ {enseignant} a {len(dates)} jour(s) de surveillance")
    print("\n3. Nombre requis d'enseignants par créneau:")
    for (date, horaire), surveillants in surveillances.items():
        nb_surveillants = len(surveillants)
        matching_seances = seances_par_creneau[seances_par_creneau['date'] == date]
        matching_seances = matching_seances[matching_seances['horaire'].apply(normaliser_horaire) == horaire]
        
        if not matching_seances.empty:
            nb_salles_necessaires = matching_seances['nombre_salles'].iloc[0]
            nb_enseignants_min = 2 * nb_salles_necessaires
            nb_enseignants_max = 3 * nb_salles_necessaires
            
            if nb_salles_necessaires > 0:
                if nb_surveillants >= nb_enseignants_min and nb_surveillants <= nb_enseignants_max:
                    print(f"   ✅ {date} {horaire}: {nb_surveillants} enseignants (besoin de {nb_enseignants_min} à {nb_enseignants_max})")
                else:
                    if nb_surveillants < nb_enseignants_min:
                        print(f"   ❌ {date} {horaire}: {nb_surveillants} enseignants (trop peu, besoin d'au moins {nb_enseignants_min})")
                    elif nb_surveillants > nb_enseignants_max:
                        print(f"   ❌ {date} {horaire}: {nb_surveillants} enseignants (trop nombreux, maximum {nb_enseignants_max})")
        else:
            print(f"   ⚠️ Aucune séance trouvée pour {date} {horaire}")
    print("\n4. Respect de la charge obligatoire maximale:")
    for enseignant, dates in creneaux_par_date.items():
        nb_surveillances = sum(len(horaires) for horaires in dates.values())
        
        matching_rows = enseignants_charges[enseignants_charges['Nom Et Prénom Enseignant'] == enseignant]
        if not matching_rows.empty:
            charge_oblig = matching_rows['charge obligatoire'].iloc[0]
            if nb_surveillances <= charge_oblig:
                print(f"   ✅ {enseignant}: {nb_surveillances}/{charge_oblig} surveillances (ne dépasse pas la charge obligatoire)")
            else:
                print(f"   ❌ {enseignant}: {nb_surveillances}/{charge_oblig} surveillances (dépasse la charge obligatoire)")
        else:
            print(f"   ⚠️ {enseignant} n'a pas de charge définie")
class NeuralOptimizer:
    def __init__(self, data_manager, solution_initiale):
        self.data_manager = data_manager
        self.solution = solution_initiale.copy()
        self.enseignants = list(data_manager.enseignants_charges['Nom Et Prénom Enseignant'])
        self.creneaux = [(date, horaire) for (date, horaire), _ in data_manager.creneaux_idx.items()]
        self.charges_obligatoires = data_manager.obtenir_charges_obligatoires()
        self.dates_seances_enseignants = data_manager.obtenir_dates_seances_enseignants()
        self.creneaux_info = data_manager.obtenir_info_creneaux()
        
    def calculer_caracteristiques(self):
        """Calcule des caractéristiques pour aider à l'optimisation"""
        n_enseignants = len(self.enseignants)
        n_creneaux = len(self.creneaux)
        compatibilite = np.zeros((n_enseignants, n_creneaux))
        
        for i, enseignant in enumerate(self.enseignants):
            dates_seances = self.dates_seances_enseignants.get(enseignant, set())
            
            for j, (date, _) in enumerate(self.creneaux):
                if date in dates_seances:
                    compatibilite[i, j] = 3.0
                dates_actuelles = set()
                for k, (d, _) in enumerate(self.creneaux):
                    if self.solution[i, k] == 1:
                        dates_actuelles.add(d)
                nouvelle_date = self.data_manager.convertir_en_datetime(date)
                dates_dt = [self.data_manager.convertir_en_datetime(d) for d in dates_actuelles]
                consecutive = True
                for d in dates_dt:
                    if abs((nouvelle_date - d).days) > 1 and abs((nouvelle_date - d).days) != len(dates_actuelles):
                        consecutive = False
                        break
                
                if consecutive:
                    compatibilite[i, j] += 2.0
        
        return compatibilite
    
    def optimiser_localement(self, iterations=100):
        """Améliore localement la solution en utilisant des concepts de réseaux neuronaux"""
        print("Optimisation locale avec approche neuronale...")
        n_enseignants = len(self.enseignants)
        n_creneaux = len(self.creneaux)
        compatibilite = self.calculer_caracteristiques()
        importance_creneaux = np.zeros(n_creneaux)
        for j, (date, horaire) in enumerate(self.creneaux):
            info = self.creneaux_info.get((date, horaire), {'min_surveillants': 0})
            nb_couvert = sum(self.solution[i, j] for i in range(n_enseignants))
            importance_creneaux[j] = max(0, info['min_surveillants'] - nb_couvert)
        if importance_creneaux.max() > 0:
            importance_creneaux = importance_creneaux / importance_creneaux.max()
        for iteration in range(iterations):
            amelioration = False
            for i in range(n_enseignants):
                charge_actuelle = sum(self.solution[i, :])
                charge_obligatoire = self.charges_obligatoires.get(self.enseignants[i], 0)
                if charge_actuelle > charge_obligatoire:
                    creneaux_assignes = [j for j in range(n_creneaux) if self.solution[i, j] == 1]
                    creneaux_assignes.sort(key=lambda j: compatibilite[i, j])
                    
                    for j in creneaux_assignes:
                        date, horaire = self.creneaux[j]
                        info = self.creneaux_info.get((date, horaire), {'min_surveillants': 0})
                        nb_couvert = sum(self.solution[k, j] for k in range(n_enseignants))
                        
                        if nb_couvert > info['min_surveillants']:
                            self.solution[i, j] = 0
                            amelioration = True
                            charge_actuelle -= 1
                            
                            if charge_actuelle <= charge_obligatoire:
                                break
                elif charge_actuelle < charge_obligatoire:
                    scores_creneaux = []
                    
                    for j in range(n_creneaux):
                        if self.solution[i, j] == 0:
                            date, horaire = self.creneaux[j]
                            info = self.creneaux_info.get((date, horaire), {'max_surveillants': 0})
                            nb_couvert = sum(self.solution[k, j] for k in range(n_enseignants))
                            if nb_couvert < info['max_surveillants']:
                                score = compatibilite[i, j] + importance_creneaux[j] * 2
                                scores_creneaux.append((j, score))
                    scores_creneaux.sort(key=lambda x: x[1], reverse=True)
                    for j, _ in scores_creneaux:
                        nouvelle_date = self.creneaux[j][0]
                        dates_actuelles = set()
                        
                        for k in range(n_creneaux):
                            if self.solution[i, k] == 1:
                                dates_actuelles.add(self.creneaux[k][0])
                        if len(dates_actuelles) < 3 or nouvelle_date in dates_actuelles:
                            self.solution[i, j] = 1
                            amelioration = True
                            charge_actuelle += 1
                            
                            if charge_actuelle >= charge_obligatoire:
                                break
                        else:
                            temp_dates = dates_actuelles.copy()
                            temp_dates.add(nouvelle_date)
                            
                            if sont_consecutives(temp_dates, self.data_manager):
                                self.solution[i, j] = 1
                                amelioration = True
                                charge_actuelle += 1
                                
                                if charge_actuelle >= charge_obligatoire:
                                    break
            if not amelioration and iteration > 10:
                print(f"Convergence après {iteration+1} itérations")
                break
            if iteration % 10 == 0:
                compatibilite = self.calculer_caracteristiques()
                for j in range(n_creneaux):
                    date, horaire = self.creneaux[j]
                    info = self.creneaux_info.get((date, horaire), {'min_surveillants': 0})
                    nb_couvert = sum(self.solution[i, j] for i in range(n_enseignants))
                    importance_creneaux[j] = max(0, info['min_surveillants'] - nb_couvert)
                
                if importance_creneaux.max() > 0:
                    importance_creneaux = importance_creneaux / importance_creneaux.max()
        
        return self.solution
def main():
    try:
        start_time = time.time()
        print("Initialisation du gestionnaire de données...")
        data_manager = DataManager().lire_donnees()
        print("\nInitialisation de l'optimiseur génétique...")
        optimiseur = GeneticOptimizer(data_manager)
        solution, logbook = optimiseur.optimiser()
        print("\nAffinage de la solution avec approche neuronale...")
        neural_optimizer = NeuralOptimizer(data_manager, solution)
        solution_finale = neural_optimizer.optimiser_localement()
        print("\nGénération des résultats...")
        result_manager = ResultManager(data_manager, solution_finale)
        surveillances = result_manager.extraire_surveillances()
        creneaux_par_date = result_manager.extraire_creneaux_par_date()
        dates_surveillance = result_manager.extraire_dates_surveillance()
        fichier_excel = result_manager.generer_excel()
        print(f"\nLe planning de surveillance optimisé a été généré dans le fichier {fichier_excel}")
        verifier_contraintes(surveillances, creneaux_par_date, dates_surveillance, data_manager)
        temps_total = time.time() - start_time
        print(f"\nTemps d'exécution total: {temps_total:.2f} secondes")
        nb_total_surveillances = sum(len(surveillants) for surveillants in surveillances.values())
        nb_enseignants = len(data_manager.enseignants_charges)
        nb_creneaux = len(surveillances)
        
        print("\nStatistiques de la solution:")
        print(f"- Nombre total de surveillances attribuées: {nb_total_surveillances}")
        print(f"- Nombre moyen de surveillances par enseignant: {nb_total_surveillances / nb_enseignants:.2f}")
        print(f"- Nombre moyen de surveillants par créneau: {nb_total_surveillances / nb_creneaux:.2f}")
        
    except FileNotFoundError as e:
        print(f"Erreur: Fichier non trouvé - {e}")
        print("Veuillez vérifier que les fichiers Excel existent dans le dossier 'resultats_surveillance'.")
    except Exception as e:
        print(f"Erreur: {e}")
        import traceback
        print(traceback.format_exc()) 

if __name__ == "__main__":
    main()